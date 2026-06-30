"""Scrape random US addresses from bestrandoms.com via direct HTTP calls.

The site's "Generate" button just reloads the page with N fresh random
addresses, which is reproducible with a plain GET to:

    https://www.bestrandoms.com/random-address-in-us?quantity=10

So instead of driving a browser, we hit that endpoint directly, parse the 10
addresses out of the returned HTML, keep only ones we have not seen before,
and append them to an Excel workbook. Repeat until interrupted.
"""

from __future__ import annotations

import time
from dataclasses import dataclass, asdict, fields
from pathlib import Path
from typing import Iterable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

URL = "https://www.bestrandoms.com/random-address-in-us"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "us_address.xlsx"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

# Maps the bold label on the page to our field name.
_LABEL_TO_FIELD = {
    "street": "street",
    "city": "city",
    "state/province/area": "state",
    "phone number": "phone",
    "zip code": "zip_code",
    "country calling code": "country_calling_code",
    "country": "country",
}


@dataclass(frozen=True)
class Address:
    street: str = ""
    city: str = ""
    state: str = ""
    phone: str = ""
    zip_code: str = ""
    country_calling_code: str = ""
    country: str = ""

    @property
    def key(self) -> tuple[str, ...]:
        """Identity used for de-duplication: a row is a duplicate only if every
        field (including phone number) matches."""
        return (
            self.street.lower().strip(),
            self.city.lower().strip(),
            self.state.lower().strip(),
            self.phone.strip(),
            self.zip_code.strip(),
            self.country_calling_code.strip(),
            self.country.lower().strip(),
        )

    def is_complete(self) -> bool:
        return bool(self.street and self.city and self.state)


# Column order in the Excel sheet.
_COLUMNS = [f.name for f in fields(Address)]
_HEADERS = {
    "street": "Street",
    "city": "City",
    "state": "State/Province/Area",
    "phone": "Phone Number",
    "zip_code": "Zip Code",
    "country_calling_code": "Country Calling Code",
    "country": "Country",
}


def parse_addresses(html: str) -> list[Address]:
    """Extract the address blocks from the page HTML."""
    soup = BeautifulSoup(html, "lxml")
    content = soup.select_one("div.content")
    if content is None:
        return []

    addresses: list[Address] = []
    for li in content.select("li.col-sm-6"):
        values: dict[str, str] = {}
        for span in li.find_all("span"):
            bold = span.find("b")
            if bold is None:
                continue
            label = bold.get_text(strip=True).rstrip(":").strip().lower()
            field = _LABEL_TO_FIELD.get(label)
            if field is None:
                continue
            # The value is the span text with the bold label removed.
            full = span.get_text(" ", strip=True)
            label_text = bold.get_text(" ", strip=True)
            value = full[len(label_text):].lstrip(": ").strip()
            values[field] = value
        addr = Address(**values)
        if addr.is_complete():
            addresses.append(addr)
    return addresses


class AddressFetcher:
    """Fetches batches of random addresses from the endpoint via HTTP GET."""

    def __init__(self, quantity: int = 10, timeout: float = 20.0) -> None:
        self.quantity = quantity
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})

    def fetch(self) -> list[Address]:
        """One 'Generate' equivalent: GET the page, return parsed addresses."""
        resp = self.session.get(
            URL, params={"quantity": self.quantity}, timeout=self.timeout
        )
        resp.raise_for_status()
        return parse_addresses(resp.text)

    def close(self) -> None:
        self.session.close()


class ExcelStore:
    """Append-only Excel store that remembers which addresses already exist."""

    def __init__(self, path: Path = DEFAULT_OUTPUT) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._seen: set[tuple[str, ...]] = set()
        self._wb, self._ws = self._open_or_create()
        self._load_seen()

    def _open_or_create(self):
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb.active
            return wb, ws
        wb = Workbook()
        ws = wb.active
        ws.title = "Addresses"
        ws.append([_HEADERS[c] for c in _COLUMNS])
        wb.save(self.path)
        return wb, ws

    def _load_seen(self) -> None:
        # Rebuild the seen-set from existing rows (skip header row).
        for row in self._ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
            data = {c: (row[i] if i < len(row) and row[i] is not None else "")
                    for i, c in enumerate(_COLUMNS)}
            self._seen.add(Address(**{k: str(v) for k, v in data.items()}).key)

    def __contains__(self, addr: Address) -> bool:
        return addr.key in self._seen

    @property
    def count(self) -> int:
        return len(self._seen)

    def add_new(self, addresses: Iterable[Address]) -> list[Address]:
        """Append addresses not already stored. Returns the ones added."""
        added: list[Address] = []
        for addr in addresses:
            if addr.key in self._seen:
                continue
            self._seen.add(addr.key)
            self._ws.append([asdict(addr)[c] for c in _COLUMNS])
            added.append(addr)
        if added:
            self._wb.save(self.path)
        return added


def scrape_forever(
    output: Path = DEFAULT_OUTPUT,
    delay: float = 1.0,
    quantity: int = 10,
) -> None:
    """Loop fetching addresses, saving new unique ones, endlessly until Ctrl+C.

    Args:
        output: workbook path.
        delay: seconds to pause between requests (be polite to the site).
        quantity: addresses requested per call (the site's quantity field).
    """
    store = ExcelStore(output)
    fetcher = AddressFetcher(quantity=quantity)
    print(f"[address-scraper] Output: {output}")
    print(f"[address-scraper] Already stored: {store.count} unique addresses")
    print("[address-scraper] Running forever — press Ctrl+C to stop.\n")

    rounds = 0
    added_total = 0
    try:
        while True:
            rounds += 1
            try:
                batch = fetcher.fetch()
            except Exception as exc:  # transient network / HTTP error
                print(f"[address-scraper] round {rounds}: error ({exc!r}); retrying...")
                time.sleep(min(10.0, delay * 3))
                continue

            added = store.add_new(batch)
            added_total += len(added)
            print(
                f"[address-scraper] round {rounds}: "
                f"{len(batch)} fetched, {len(added)} new "
                f"(total stored: {store.count})"
            )
            time.sleep(delay)
    except KeyboardInterrupt:
        print("\n[address-scraper] Stopped by user.")
    finally:
        fetcher.close()
        print(
            f"[address-scraper] Done. Added {added_total} new this session; "
            f"{store.count} unique addresses in {output}"
        )
